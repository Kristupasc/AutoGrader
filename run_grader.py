import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

from grader import Grader


class RunGrader:
    """
    Class responsible to run the grader for all exercises in the exam. It uses the exam class created
    by the grader to grade each exercise.
    :param data_file: The path to the data Excel file containing the exam data.
    :param exam_class: The class that will be used to grade the exam (written by the grader)
    :param questions: The list of questions that will be graded. For example: [5,6,7,8,9,10]
    """

    def __init__(self, data_file, exam_class, questions):
        self.data_file = data_file
        self.exam_class = exam_class
        self.questions = questions
        self.candidates = {}  # candidate id: [[question_number, answer]]
        self.candidate_results = {}  # candidate id: [[question_number, points]]
        print("Exam grader loaded. Ready to grade exams.")

    def begin(self):
        """
        Begins the grading process. It reads the data file and grades each exam using the exam class.
        """
        print("Reading data file...")
        wb = openpyxl.load_workbook(self.data_file)
        sheet = wb.active
        candidates = {}  # candidate id: [[question_number, answer]]
        keys = {}  # Column name: index
        for row in sheet.iter_rows(values_only=True):
            if not keys:
                for i in range(len(row)):
                    keys[row[i]] = i
            else:
                candidate_id = row[keys["Candidate ID"]]
                question_number = row[keys["Question number"]]
                answer = row[keys["Variable value"]]
                if candidate_id not in candidates:
                    candidates[candidate_id] = []
                if row[keys["Variable name"]] != "RESPONSE":
                    continue
                if not isinstance(answer, str) or len(answer) < 20 or not "def" in answer:
                    if question_number in self.questions:
                        candidates[candidate_id].append([question_number, None])
                    continue
                if question_number in self.questions:
                    candidates[candidate_id].append([question_number, answer])

        self.candidates = candidates

    def grade(self):
        print("Starting to grade the candidates.")
        for candidate_id in self.candidates:
            print("Grading candidate: ", candidate_id)
            if candidate_id not in self.candidate_results:
                self.candidate_results[candidate_id] = []
            exam = self.exam_class()
            for question in self.candidates[candidate_id]:
                print("Grading question: ", question[1])
                if question[1] is None:
                    self.candidate_results[candidate_id].append([question[0], 0, "No code provided (0 points)"])
                    continue
                (executed, error) = self.execute_test(question[0], exam, question[1])
                if not executed:
                    self.candidate_results[candidate_id].append(
                        [question[0], -1, f"Error in code: ({error})"])
                    continue
                self.candidate_results[candidate_id].append(
                    [question[0], exam.points[question[0]], exam.feedback[question[0]]])

        print("Grading finished.")

    def execute_test(self, question_id, exam, source):
        """
        Executes a test for a specific question.
        :param question_id: The question number to be tested.
        """
        grader = Grader(source)
        if not grader.initialized:
            return False, grader.error
        if question_id == 1:
            exam.question_1(grader)
        elif question_id == 2:
            exam.question_2(grader)
        elif question_id == 3:
            exam.question_3(grader)
        elif question_id == 4:
            exam.question_4(grader)
        elif question_id == 5:
            exam.question_5(grader)
        elif question_id == 6:
            exam.question_6(grader)
        elif question_id == 7:
            exam.question_7(grader)
        elif question_id == 8:
            exam.question_8(grader)
        elif question_id == 9:
            exam.question_9(grader)
        elif question_id == 10:
            exam.question_10(grader)

        return True, ""

    def process_results(self):
        """
        Processes the results of the grading process.
        """
        sum_questions = {i: 0 for i in self.questions}
        ungraded = {i: 0 for i in self.questions}
        graded = {i: 0 for i in self.questions}
        compile_errors = {i: 0 for i in self.questions}
        runtime_errors = {i: 0 for i in self.questions}
        for candidate in self.candidate_results:
            for question in self.candidate_results[candidate]:
                if question[1] == -1:
                    if "indent" in question[2] or "syntax" in question[2]:
                        compile_errors[question[0]] += 1
                    else:
                        runtime_errors[question[0]] += 1
                    ungraded[question[0]] += 1
                    continue
                graded[question[0]] += 1
                sum_questions[question[0]] += question[1]

        for question in self.questions:
            print(f"Average of question {str(question)}: ",
                  sum_questions[question] / (graded[question]))
            print("Ungraded: ", ungraded[question])
            print("Graded: ", graded[question])
            print("Average grading: ", (graded[question] / (graded[question] + ungraded[question])) * 100, "%")
            print("Compile errors: ", compile_errors[question])
            print("Runtime errors: ", runtime_errors[question])

        print("total average points: ", sum([sum_questions[i] for i in sum_questions])/sum([graded[i] for i in graded]))

    def export_results(self, filename="data/results.xlsx"):
        """
        Exports the results to an Excel file.
        """
        # create excel file
        wb = openpyxl.Workbook()
        sheet = wb.active

        # columns will be: Candidate ID, Question n points, Question n code, ..., Total points
        columns_points_code = []
        for i in self.questions:
            columns_points_code.append(f"Question {i} points")
            columns_points_code.append(f"Question {i} code")
            columns_points_code.append(f"Question {i} feedback")
        header = ["Candidate ID"] + columns_points_code + ["Total points"]
        sheet.append(header)

        for candidate in self.candidate_results:
            row = [candidate]
            total = 0
            for question_id in self.questions:
                points = "Not graded"
                code = "Not provided"
                feedback = "No feedback"

                # Get the candidate's answer for this question if it exists
                for q in self.candidates[candidate]:
                    if q[0] == question_id:
                        if q[1] is None:
                            code = "Not provided"
                        else:
                            code = q[1]
                        break

                # Get the candidate's score for this question if it exists
                for q in self.candidate_results[candidate]:
                    if q[0] == question_id:
                        feedback = str(q[2])
                        if q[1] != -1:
                            points = q[1]
                            total += q[1]
                        break

                row.append(points)
                row.append(str(code))
                row.append(feedback)

            row.append(total)
            sheet.append(row)

        # Apply autofilter to the header row
        sheet.auto_filter.ref = sheet.dimensions

        # Define alignment for feedback cells
        wrap_center_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Apply the alignment and set the column width for all feedback columns
        feedback_column_width = 25  # Set the desired width for feedback columns
        for col in range(4, len(header), 3):  # feedback columns are every third column starting from index 4
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = feedback_column_width
            for row in range(2, sheet.max_row + 1):  # start from the second row to skip the header
                sheet.cell(row=row, column=col).alignment = wrap_center_alignment

        try:
            wb.save(filename)
        except PermissionError:
            print("\n\n\n\nERROR:Please close the file and try again.")
            return
